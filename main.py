import time
import requests
import os
import shutil
from pprint import pprint
import awsrefs as aws
import settings
import xlsxwriter
import openpyxl as xl
import collections
from concurrent.futures import ThreadPoolExecutor

FAMILY_PER_TAB = settings.FAMILY_PER_TAB
LOOKUP_CODE = settings.LOOKUP_CODE
RI_INPUT_TEMPLATE = settings.RI_INPUT_TEMPLATE
PLAN_TYPE = settings.PLAN_TYPE
INSTANCE_FAMILY = settings.INSTANCE_FAMILY
PLAN_LENGTH = settings.PLAN_LENGTH
PLAN_COMMIT = settings.PLAN_COMMIT
REGIONS = settings.REGIONS
OSES = settings.OSES
TENANCY = settings.TENANCY

#FIXED URL
BASE_URL = "https://view-publish.us-west-2.prod.pricing.aws.a2z.com/pricing/2.0/meteredUnitMaps/computesavingsplan/USD/current/"
MID_URLS = {'compute':"compute-savings-plan-ec2",'ec2':"instance-savings-plan-ec2"}
END_URL = "index.json?timestamp="


def get_terms():
    '''Manipulates the PLAN_LENGTH and PLAN_COMMIT inputs'''
    global SP_TERMS
    terms = []
    for length in PLAN_LENGTH:
        for commit in PLAN_COMMIT:
            terms.append("{}{}".format(int(length),commit))
    SP_TERMS = terms
    

def construct_urls():
    '''Builds URL set from input'''
    URLS = []
    TIMESTAMP = str(int(time.time()))
    START_URL = "{}{}".format(BASE_URL, MID_URLS[PLAN_TYPE.lower()])
    for region in REGIONS:
        for k,v in aws.regions.items():
            if region == k:
                region = v

        for os in OSES:
            for terms in SP_TERMS:
                terms = list(terms)
                for k,v in aws.commit.items():
                    if terms[1] == k:
                        term = "{} year/{}".format(str(terms[0]),v)
                
                for tenancy in TENANCY:
                    if PLAN_TYPE.lower() == 'compute':
                        URLS.append("{}/{}/{}/{}/{}/{}{}".format(START_URL, term, region, os, tenancy, END_URL, TIMESTAMP))
                    if PLAN_TYPE.lower() == 'ec2':
                        for instance in INSTANCE_FAMILY: 
                            URLS.append("{}/{}/{}/{}/{}/{}/{}{}".format(START_URL, term, instance, region, os, tenancy, END_URL, TIMESTAMP))

    print("Working on {} URLS".format(len(URLS)))
    return URLS


def get_json(in_url):
    ''' Parse json to ordered dict '''
    response = requests.get(in_url)
    if response.status_code != 200:
        print("ERROR : Could not connect to {}".format(in_url))
    working_data = response.json()
    for i in working_data['regions'].values():
        for k, v in i.items():
            entry = {}
            rate = v
            tenancy = rate['ec2:Tenancy']
            instance = rate['ec2:InstanceType']
            sp_region = rate['ec2:Location']
            for k,v in aws.regions.items():
                if sp_region == v:
                    sp_region, sp_region_code = v, k

            operatingsystem = rate['ec2:OperatingSystem']
            for k,v in aws.os.items():
                if operatingsystem == v:
                    operatingsystem = k
        
            commityear = rate['LeaseContractLength']
            commitamount = rate['PurchaseOption']
            for k,v in aws.commit.items():
                if commitamount  == v:
                    commitamount  = k

            commitcode = commityear + commitamount

            sprate = rate['price']
            odrate = rate['ec2:PricePerUnit']
            spcode = "{}-{}-{}-{}-{}".format(instance, sp_region, operatingsystem, tenancy, commitcode)
            savingper = ((float(odrate)-float(sprate))/float(odrate))*100

            entry_key = instance+spcode
            entry = {"instance": instance,
                    "region": sp_region,
                    "regioncode": sp_region_code,
                    "os": operatingsystem,
                    "tenancy": tenancy,
                    "commitcode": commitcode,
                    "odrate": "{:0.4f}".format(float(odrate)),
                    "sprate": sprate,
                    "savingper": "{:0.2f}".format(savingper),
                    }
            if LOOKUP_CODE == True:
                entry.update({"spcode" : spcode})

            # Tabulate Excel if True
            if FAMILY_PER_TAB == True:
                xls_tab = instance[0]
            else:
                xls_tab = 'SavingsPlans'

            if xls_tab not in response_dict:
                response_dict[xls_tab] = collections.OrderedDict()
                response_dict[xls_tab][entry_key] = entry
            else:
                response_dict[xls_tab][entry_key] = entry


def xlwriter(response_dict):
    ''' Write to XSLX '''
    workbook = xlsxwriter.Workbook('temp.xlsx')
    bold = workbook.add_format({'bold': True})
    money = workbook.add_format({'num_format': '$#,##0.0000'})

    for k, v in response_dict.items():
        # Create Worksheet
        worksheet = workbook.add_worksheet(k)

        # Add Header Row
        header_row = ["instance", "region", "regioncode", "os", "tenancy", "commitcode", "lookup", "odrate", "sprate", "% Saving"]
        if LOOKUP_CODE == False:
            header_row.remove("lookup")
        
        for i, e in enumerate(header_row):
            worksheet.write(0, i, e, bold)

        # Iterate Rows
        row = 1
        for k, v in v.items():
            
            content_row = [v["instance"], v["region"], v["regioncode"], v["os"], v["tenancy"], v["commitcode"]]
            if LOOKUP_CODE == True:
                content_row.append(v["spcode"])

            for i, e in enumerate(content_row):
                worksheet.write(row, i, e)
                col = i

            worksheet.write_number(row, col+1, float(v["odrate"]), money)
            worksheet.write_number(row, col+2, float(v["sprate"]), money)
            worksheet.write_number(row, col+3, float(v["savingper"]))
            row += 1
    workbook.close()

def merge_spreadsheets():
    xls_file = "{}-savings-plans.xlsx".format(PLAN_TYPE.lower())
    
    if RI_INPUT_TEMPLATE == True:
        if os.path.exists(xls_file):
            os.remove(xls_file)
        shutil.copyfile('template.xlsx', xls_file)
        in_xls = 'temp.xlsx'
        out_xls = xls_file
        wb1 = xl.load_workbook(filename=in_xls)
        ws1 = wb1.worksheets[0]
        wb2 = xl.load_workbook(filename=out_xls)
        ws2 = wb2['SavingsPlans']
        for row in ws1:
            for cell in row:
                ws2[cell.coordinate].value = cell.value
        wb2.save(out_xls)
    else:
        os.rename('temp.xlsx', xls_file)
        

def main():
    ''' Main entry point of the app '''
    get_terms()
    print("\nRegions - {}".format(', '.join(map(str, REGIONS))))
    print("OS - {}".format(', '.join(map(str, OSES))))
    print("Tenancy - {}".format(', '.join(map(str, TENANCY))))
    print("Commitment Types - {}".format(', '.join(map(str, SP_TERMS))))
    working_urls = construct_urls()

    with ThreadPoolExecutor() as executor:
        executor.map(get_json, working_urls, timeout=30)

    xlwriter(response_dict)
    merge_spreadsheets()


if __name__ == "__main__":
    ''' This is executed when run from the command line '''
    start = time.time()
    response_dict = collections.OrderedDict()
    main()
    stop = time.time()
    print("\nRuntime {:0.2f}s\n".format(stop-start))